import tempfile
from pathlib import Path
import pytest
import openpyxl
from app.core.excel_handler import ExcelHandler, MatchMode


def _make_excel(path: Path, headers: list, rows: list[list]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    wb.save(path)
    return path


def _make_lrmx(path: Path, fields: dict[str, str]) -> Path:
    import xml.etree.ElementTree as ET
    root = ET.Element('Person')
    for k, v in fields.items():
        elem = ET.SubElement(root, k)
        elem.text = v
    tree = ET.ElementTree(root)
    ET.indent(tree, space='    ')
    tree.write(str(path), encoding='UTF-8', xml_declaration=True)
    return path


@pytest.fixture
def temp_dir():
    with tempfile.TemporaryDirectory() as d:
        yield Path(d)


class TestExportToExcel:
    def test_basic_export(self, temp_dir):
        excel = _make_excel(temp_dir / 'test.xlsx',
            ['身份证号', '姓名', '出生年月'],
            [['110101199001011234', '张三', '']])
        lrmx = _make_lrmx(temp_dir / 'test.lrmx',
            {'XingMing': '张三', 'ShenFenZheng': '110101199001011234', 'ChuShengNianYue': '199001'})

        handler = ExcelHandler(excel, [lrmx], MatchMode.ID_CARD)
        handler.export_to_excel(
            field_mapping={'ChuShengNianYue': '出生年月'},
            fields_to_write=['ChuShengNianYue'],
            header_row=1,
            match_excel_col_for_id='身份证号',
        )

        wb = openpyxl.load_workbook(excel)
        ws = wb.active
        assert ws.cell(row=2, column=3).value == '199001'

    def test_unmatched_skipped(self, temp_dir):
        excel = _make_excel(temp_dir / 'test.xlsx',
            ['身份证号', '姓名'], [['110101199001011234', '张三']])
        lrmx = _make_lrmx(temp_dir / 'test.lrmx',
            {'XingMing': '李四', 'ShenFenZheng': '999999999999999999'})

        handler = ExcelHandler(excel, [lrmx], MatchMode.ID_CARD)
        logs = handler.export_to_excel(
            field_mapping={}, fields_to_write=[],
            header_row=1, match_excel_col_for_id='身份证号',
        )
        assert any('未在名册中找到匹配记录' in m for m in logs)

    def test_new_columns_appended(self, temp_dir):
        excel = _make_excel(temp_dir / 'test.xlsx',
            ['身份证号', '姓名'], [['110101199001011234', '张三']])
        lrmx = _make_lrmx(temp_dir / 'test.lrmx',
            {'XingMing': '张三', 'ShenFenZheng': '110101199001011234', 'XueLi': '本科'})

        handler = ExcelHandler(excel, [lrmx], MatchMode.ID_CARD)
        handler.export_to_excel(
            field_mapping={'XueLi': '最高学历'},
            fields_to_write=['XueLi'],
            header_row=1,
            match_excel_col_for_id='身份证号',
        )

        wb = openpyxl.load_workbook(excel)
        ws = wb.active
        assert ws.cell(row=1, column=3).value == '最高学历'
        assert ws.cell(row=2, column=3).value == '本科'

    def test_with_converter(self, temp_dir):
        excel = _make_excel(temp_dir / 'test.xlsx',
            ['身份证号', '姓名', '出生年月'],
            [['110101199001011234', '张三', '']])
        lrmx = _make_lrmx(temp_dir / 'test.lrmx',
            {'XingMing': '张三', 'ShenFenZheng': '110101199001011234', 'ChuShengNianYue': '199001'})

        from app.core.converters import BUILTIN_CONVERTERS
        date_code = BUILTIN_CONVERTERS[0]['code']

        handler = ExcelHandler(excel, [lrmx], MatchMode.ID_CARD)
        handler.export_to_excel(
            field_mapping={'ChuShengNianYue': '出生年月'},
            fields_to_write=['ChuShengNianYue'],
            converters={'ChuShengNianYue': date_code},
            header_row=1,
            match_excel_col_for_id='身份证号',
        )

        wb = openpyxl.load_workbook(excel)
        ws = wb.active
        assert ws.cell(row=2, column=3).value == '1990.01'

    def test_backup_created(self, temp_dir):
        excel = _make_excel(temp_dir / 'test.xlsx',
            ['身份证号', '姓名'], [['110101199001011234', '张三']])
        lrmx = _make_lrmx(temp_dir / 'test.lrmx',
            {'XingMing': '张三', 'ShenFenZheng': '110101199001011234', 'XueLi': '本科'})

        handler = ExcelHandler(excel, [lrmx], MatchMode.ID_CARD)
        handler.export_to_excel(
            field_mapping={'XueLi': '学历'}, fields_to_write=['XueLi'],
            header_row=1, match_excel_col_for_id='身份证号',
        )
        assert (temp_dir / 'test.xlsx.bak').exists()
