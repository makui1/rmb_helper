from pathlib import Path
import tempfile
import openpyxl
from app.core.verify_handler import PersonResult, FieldResult
from app.core.result_exporter import export_excel, export_html


def _sample_results() -> list[PersonResult]:
    ok = PersonResult(
        name='张三', lrmx_path='a.lrmx', status='ok',
        fields=[
            FieldResult(field='XingMing', excel_val='张三', lrmx_val='张三', match=True),
            FieldResult(field='ShenFenZheng', excel_val='110101199001011234',
                        lrmx_val='110101199001011234', match=True),
        ],
    )
    diff = PersonResult(
        name='李四', lrmx_path='b.lrmx', status='diff',
        fields=[
            FieldResult(field='ShenFenZheng', excel_val='110101199001015678',
                        lrmx_val='110101199001015678', match=True),
            FieldResult(field='ChuShengNianYue', excel_val='1990-01',
                        lrmx_val='1990.01', match=False),
        ],
    )
    nf = PersonResult(name='王五', lrmx_path='c.lrmx', status='not_found')
    err = PersonResult(name='赵六', lrmx_path='d.lrmx', status='error',
                       error_msg='文件损坏')
    return [ok, diff, nf, err]


def test_excel_sheet1_headers_and_row_count():
    with tempfile.TemporaryDirectory() as d:
        path = Path(d) / 'out.xlsx'
        export_excel(_sample_results(), path)
        wb = openpyxl.load_workbook(str(path))
        ws = wb['人员汇总']
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ('姓名', '身份证', '核验状态', '差异字段数', '差异字段', '错误信息')
        assert len(rows) == 5  # 1 header + 4 results


def test_excel_sheet1_status_labels():
    with tempfile.TemporaryDirectory() as d:
        path = Path(d) / 'out.xlsx'
        export_excel(_sample_results(), path)
        wb = openpyxl.load_workbook(str(path))
        ws = wb['人员汇总']
        rows = list(ws.iter_rows(values_only=True))
        assert rows[1][2] == '一致'
        assert rows[2][2] == '有差异'
        assert rows[2][3] == 1        # 1 差异字段
        assert rows[2][4] == 'ChuShengNianYue'
        assert rows[3][2] == '名册无此人'
        assert rows[4][2] == '错误'
        assert rows[4][5] == '文件损坏'


def test_excel_sheet2_field_rows():
    with tempfile.TemporaryDirectory() as d:
        path = Path(d) / 'out.xlsx'
        export_excel(_sample_results(), path)
        wb = openpyxl.load_workbook(str(path))
        ws = wb['字段明细']
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ('姓名', '身份证', '字段', '名册值', '任免表值', '是否一致')
        field_col = [r[2] for r in rows[1:]]
        # ok 的 2 个字段 + diff 的 2 个字段 + not_found 1行 + error 1行
        assert len(field_col) == 6
        assert 'ChuShengNianYue' in field_col
        assert '—' in field_col   # not_found / error 占位行


def test_excel_sheet2_match_symbols():
    with tempfile.TemporaryDirectory() as d:
        path = Path(d) / 'out.xlsx'
        export_excel(_sample_results(), path)
        wb = openpyxl.load_workbook(str(path))
        ws = wb['字段明细']
        rows = list(ws.iter_rows(values_only=True))
        match_col = [r[5] for r in rows[1:] if r[5] in ('✓', '✗')]
        assert '✓' in match_col
        assert '✗' in match_col


def test_html_creates_valid_file():
    with tempfile.TemporaryDirectory() as d:
        path = Path(d) / 'out.html'
        export_html(_sample_results(), path, config_summary='测试摘要')
        content = path.read_text(encoding='utf-8')
        assert content.startswith('<!DOCTYPE html>')
        assert '测试摘要' in content
        assert '张三' in content
        assert '李四' in content
        assert 'class="badge ok"' in content
        assert 'class="badge diff"' in content
        assert 'class="badge not_found"' in content
        assert 'class="badge error"' in content


def test_html_diff_highlighting():
    results = [PersonResult(
        name='测试人', lrmx_path='x.lrmx', status='diff',
        fields=[FieldResult(
            field='ChuShengNianYue',
            excel_val='1990-01',
            lrmx_val='1990.01',
            match=False,
        )],
    )]
    with tempfile.TemporaryDirectory() as d:
        path = Path(d) / 'out.html'
        export_html(results, path)
        content = path.read_text(encoding='utf-8')
        assert 'class="del"' in content
        assert 'class="ins"' in content


def test_html_diff_open_by_default():
    """'有差异'的 details 块默认 open，'一致'的不展开。"""
    results = _sample_results()
    with tempfile.TemporaryDirectory() as d:
        path = Path(d) / 'out.html'
        export_html(results, path)
        content = path.read_text(encoding='utf-8')
        # diff entry should have <details open>
        assert '<details open>' in content
        # ok entry should just be <details>
        assert '<details>' in content


def test_html_not_found_and_error_note():
    results = [
        PersonResult(name='王五', lrmx_path='c.lrmx', status='not_found'),
        PersonResult(name='赵六', lrmx_path='d.lrmx', status='error', error_msg='文件损坏'),
    ]
    with tempfile.TemporaryDirectory() as d:
        path = Path(d) / 'out.html'
        export_html(results, path)
        content = path.read_text(encoding='utf-8')
        assert '未找到匹配记录' in content
        assert '文件损坏' in content
