"""Export verification results to Excel or self-contained HTML."""
from __future__ import annotations

import html as _html
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from app.core.verify_handler import PersonResult, char_diff_html

# ── shared constants ──────────────────────────────────────────────────────────

_STATUS_LABELS: dict[str, str] = {
    'ok':        '一致',
    'diff':      '有差异',
    'not_found': '名册无此人',
    'error':     '错误',
}

_STATUS_FILL: dict[str, PatternFill] = {
    'ok':        PatternFill('solid', fgColor='E8F5EC'),
    'diff':      PatternFill('solid', fgColor='FDEAEA'),
    'not_found': PatternFill('solid', fgColor='FFF3E0'),
    'error':     PatternFill('solid', fgColor='F5F5F5'),
}

_DEL_FILL = PatternFill('solid', fgColor='FDEAEA')
_INS_FILL = PatternFill('solid', fgColor='E8F5EC')

_BOLD = Font(bold=True)


def _id_val(result: PersonResult) -> str:
    """Return the Excel value for ShenFenZheng, or '' if not mapped."""
    for fr in result.fields:
        if fr.field == 'ShenFenZheng':
            return fr.excel_val
    return ''


# ── Excel export ──────────────────────────────────────────────────────────────

def export_excel(results: list[PersonResult], path: Path) -> None:
    """Write a dual-sheet workbook. Raises on failure."""
    wb = openpyxl.Workbook()

    # ── Sheet1: 人员汇总 ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = '人员汇总'
    headers1 = ['姓名', '身份证', '核验状态', '差异字段数', '差异字段', '错误信息']
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = _BOLD

    for r in results:
        status = r.status if r.status in _STATUS_LABELS else 'error'
        if r.fields:
            diff_n: int | str = sum(1 for f in r.fields if not f.match)
            diff_names = ', '.join(f.field for f in r.fields if not f.match)
        else:
            diff_n = ''
            diff_names = ''
        ws1.append([
            r.name,
            _id_val(r),
            _STATUS_LABELS[status],
            diff_n,
            diff_names,
            r.error_msg,
        ])
        ws1.cell(ws1.max_row, 3).fill = _STATUS_FILL[status]

    # ── Sheet2: 字段明细 ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet('字段明细')
    headers2 = ['姓名', '身份证', '字段', '名册值', '任免表值', '是否一致']
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = _BOLD

    for r in results:
        id_ = _id_val(r)
        status = r.status if r.status in _STATUS_LABELS else 'error'
        if status in ('ok', 'diff') and r.fields:
            for fr in r.fields:
                ws2.append([
                    r.name, id_, fr.field,
                    fr.excel_val, fr.lrmx_val,
                    '✓' if fr.match else '✗',
                ])
                if not fr.match:
                    ws2.cell(ws2.max_row, 4).fill = _DEL_FILL
                    ws2.cell(ws2.max_row, 5).fill = _INS_FILL
        else:
            note = r.error_msg if status == 'error' else _STATUS_LABELS[status]
            ws2.append([r.name, id_, '—', note, '', ''])

    wb.save(path)


# ── HTML export ───────────────────────────────────────────────────────────────

_HTML_CSS = """\
body{font-family:system-ui,sans-serif;margin:24px;color:#333;background:#F8F5F0}
.header{margin-bottom:16px}
.meta{font-size:12px;color:#888880;margin-bottom:8px}
.counts{display:flex;gap:12px;flex-wrap:wrap}
.badge{font-size:12px;font-weight:600;padding:2px 8px;border-radius:4px;white-space:nowrap}
.badge.ok{color:#1E7A3A;background:#E8F5EC}
.badge.diff{color:#B02020;background:#FDEAEA}
.badge.not_found{color:#C07030;background:#FFF3E0}
.badge.error{color:#555;background:#F5F5F5}
details{background:#fff;border-radius:6px;margin-bottom:6px;border:1px solid #E8E4DE}
summary{padding:8px 12px;cursor:pointer;display:flex;align-items:center;gap:8px;
        list-style:none;user-select:none}
summary::-webkit-details-marker{display:none}
.arrow{font-size:10px;color:#888}
.name{font-weight:500;flex:1}
table{width:100%;border-collapse:collapse;font-size:12px}
th{background:#F5F2EE;color:#888880;font-weight:500;padding:5px 10px;text-align:left}
td{padding:5px 10px;border-top:1px solid #F0EDEA;vertical-align:top;word-break:break-all}
.field-col{color:#888880;width:160px;word-break:normal}
.del{background:#FDEAEA;color:#B02020;border-radius:2px;padding:0 2px}
.ins{background:#E8F5EC;color:#1E7A3A;border-radius:2px;padding:0 2px}
.same{color:#AAAAAA;font-style:italic}
.note{padding:8px 12px;font-size:12px;color:#888}
"""


def export_html(
    results: list[PersonResult],
    path: Path,
    config_summary: str = '',
) -> None:
    """Write a self-contained HTML file reproducing the verify-tab diff view."""
    counts: dict[str, int] = {'ok': 0, 'diff': 0, 'not_found': 0, 'error': 0}
    for r in results:
        counts[r.status if r.status in counts else 'error'] += 1

    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    meta_text = f'导出时间：{now}'
    if config_summary:
        meta_text += f'　·　{config_summary}'

    parts: list[str] = [
        '<!DOCTYPE html><html lang="zh"><head>'
        '<meta charset="utf-8"><title>核验结果</title>'
        f'<style>{_HTML_CSS}</style></head><body>',
        '<div class="header">',
        f'<div class="meta">{_html.escape(meta_text)}</div>',
        '<div class="counts">',
    ]
    for key, label in _STATUS_LABELS.items():
        parts.append(
            f'<span class="badge {key}">{counts[key]} {label}</span>'
        )
    parts.append('</div></div>')

    for r in results:
        status = r.status if r.status in _STATUS_LABELS else 'error'
        label = _STATUS_LABELS[status]
        open_attr = ' open' if status == 'diff' else ''
        name_escaped = _html.escape(r.name or r.lrmx_path)

        parts.append(f'<details{open_attr}>')
        parts.append(
            f'<summary>'
            f'<span class="arrow">▶</span>'
            f'<span class="name">{name_escaped}</span>'
            f'<span class="badge {status}">{label}</span>'
            f'</summary>'
        )

        if status in ('ok', 'diff') and r.fields:
            parts.append(
                '<table><tr><th>字段</th><th>名册值</th><th>任免表值</th></tr>'
            )
            for fr in r.fields:
                field_esc = _html.escape(fr.field)
                if fr.match:
                    val = _html.escape(fr.lrmx_val or '')
                    parts.append(
                        f'<tr><td class="field-col">{field_esc}</td>'
                        f'<td colspan="2" class="same">{val}（一致）</td></tr>'
                    )
                else:
                    a_html, b_html = char_diff_html(fr.excel_val, fr.lrmx_val)
                    parts.append(
                        f'<tr><td class="field-col">{field_esc}</td>'
                        f'<td>{a_html}</td><td>{b_html}</td></tr>'
                    )
            parts.append('</table>')
        elif status == 'not_found':
            parts.append('<div class="note">该人员在名册中未找到匹配记录。</div>')
        else:  # error
            parts.append(
                f'<div class="note">错误：{_html.escape(r.error_msg)}</div>'
            )

        parts.append('</details>')

    parts.append('</body></html>')
    path.write_text(''.join(parts), encoding='utf-8')
