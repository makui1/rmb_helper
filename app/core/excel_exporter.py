import re
import openpyxl
from datetime import date
from pathlib import Path
import sys


def get_template_path() -> Path:
    """返回内置模板路径，开发环境和 PyInstaller 打包后均适用。"""
    if hasattr(sys, '_MEIPASS'):
        p = Path(sys._MEIPASS) / 'resources' / 'family_template.xlsx'
    else:
        p = Path(__file__).parent.parent / 'resources' / 'family_template.xlsx'
    if not p.exists():
        raise FileNotFoundError(
            f'找不到内置模板文件，请将 family_template.xlsx 放至：{p}'
        )
    return p

SPECIAL_CHENGWEI = {
    '父亲', '母亲',
    '祖父', '爷爷',
    '祖母', '奶奶',
    '外祖父', '外祖母',
    '孙子', '孙女',
}

_PLACEHOLDER_RE = re.compile(r'\{\{(\w+(?:\.\w+)?)\}\}')
_BIRTH_YM_RE    = re.compile(r'\d{4}\.\d{1,2}')


def fmt_birth(raw: str) -> str:
    """'196811' → '1968.11' (ensure month is 2-digit)"""
    raw = (raw or '').strip()
    if re.fullmatch(r'\d{6}', raw):
        year, month = raw[:4], raw[4:]
        return f'{year}.{int(month):02d}'
    return raw


def birth_with_age(birth_ym: str) -> str:
    """'1965.10' → '1965.10\\n（61岁）' using current year."""
    if not birth_ym:
        return birth_ym
    m = re.match(r'(\d{4})', birth_ym)
    if not m:
        return birth_ym
    age = date.today().year - int(m.group(1))
    return f'{birth_ym}\n（{age}岁）'


def normalize_birth(value: str) -> str:
    """Normalize birth date to yyyy.MM format.
    Handles both '196811' (raw) and '1956.4' (malformed) → '1968.11' / '1956.04'
    """
    value = (value or '').strip()
    if not value:
        return ''
    if re.fullmatch(r'\d{6}', value):
        return fmt_birth(value)
    m = re.match(r'(\d{4})\.(\d{1,2})', value)
    if m:
        year, month = m.group(1), m.group(2)
        return f'{year}.{int(month):02d}'
    return value


class ExcelExporter:
    def __init__(self, template_path: Path = get_template_path()):
        self._template = template_path
        # {(sheet_title, coordinate): field_name}
        self._field_map: dict[tuple[str, str], str] = {}
        self._name_coord:  tuple[str, str] | None = None
        self._birth_coord: tuple[str, str] | None = None
        self._scan_template()

    def _scan_template(self):
        wb = openpyxl.load_workbook(self._template)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    m = _PLACEHOLDER_RE.search(cell.value)
                    if not m:
                        continue
                    field = m.group(1)
                    key   = (ws.title, cell.coordinate)
                    self._field_map[key] = field
                    if '.' not in field:
                        if field == 'XingMing' and self._name_coord is None:
                            self._name_coord = key
                        elif field == 'ChuShengNianYue' and self._birth_coord is None:
                            self._birth_coord = key
        wb.close()

    def scan_output_dir(self, output_dir: Path) -> dict[tuple[str, str], Path]:
        """Return {(name, birth_ym): path} for existing xlsx in output_dir.
        Normalizes birth dates to yyyy.MM format.
        """
        result: dict[tuple[str, str], Path] = {}
        if not output_dir.is_dir():
            return result
        for p in output_dir.glob('*.xlsx'):
            try:
                wb   = openpyxl.load_workbook(p, read_only=True, data_only=True)
                name = birth = ''
                if self._name_coord:
                    sh, coord = self._name_coord
                    ws = wb[sh] if sh in wb.sheetnames else wb.active
                    v  = ws[coord].value
                    name = str(v).strip() if v else ''
                if self._birth_coord:
                    sh, coord = self._birth_coord
                    ws = wb[sh] if sh in wb.sheetnames else wb.active
                    v  = ws[coord].value
                    raw = str(v) if v else ''
                    birth = normalize_birth(raw)
                wb.close()
                if name and birth:
                    result[(name, birth)] = p
            except Exception:
                pass
        return result

    def export(
        self,
        lrmx_dict:    dict[str, str],
        members:      list[dict[str, str]],
        output_path:  Path,
        existing_map: dict[tuple[str, str], Path],
        on_exists:    str = 'skip',
        fix_birth:    bool = False,
    ) -> tuple[str, str]:
        """Export to xlsx.

        Args:
            fix_birth: If True and updating, also rewrite ChuShengNianYue to standard format

        Returns (status, stem):
          'created' — new file written
          'updated' — existing file basic-info refreshed
          'skip'    — existing file unchanged
        """
        name      = lrmx_dict.get('XingMing', '')
        birth_raw = lrmx_dict.get('ChuShengNianYue', '')
        birth_ym  = fmt_birth(birth_raw)
        stem      = output_path.stem

        existing = existing_map.get((name, birth_ym))

        if existing is not None:
            if on_exists == 'skip':
                return 'skip', stem
            wb          = openpyxl.load_workbook(existing)
            save_path   = existing
            update_only = True
        else:
            wb          = openpyxl.load_workbook(self._template)
            save_path   = output_path
            update_only = False

        n_group = [m for m in members if m.get('ChengWei') in SPECIAL_CHENGWEI]
        m_group = [m for m in members if m.get('ChengWei') not in SPECIAL_CHENGWEI]

        for (sh, coord), field in self._field_map.items():
            is_member = '.' in field
            if update_only and is_member:
                continue  # preserve manually-edited family data

            ws   = wb[sh] if sh in wb.sheetnames else wb.active
            cell = ws[coord]

            if is_member:
                prefix, attr = field.split('.', 1)
                if prefix.startswith('m'):
                    idx    = int(prefix[1:])
                    member = m_group[idx] if idx < len(m_group) else None
                elif prefix.startswith('n'):
                    idx    = int(prefix[1:])
                    member = n_group[idx] if idx < len(n_group) else None
                else:
                    member = None

                if member and attr == 'ChuShengNianYue':
                    cell.value = fmt_birth(member.get('ChuShengRiQi', '')) or None
                elif member:
                    cell.value = member.get(attr, '') or None
                else:
                    cell.value = None  # truly empty — ISBLANK() returns True
            else:
                if field == 'ChuShengNianYue':
                    bym = normalize_birth(birth_ym) if (update_only and fix_birth) else birth_ym
                    cell.value = birth_with_age(bym)
                    # enable text wrap so the age line displays correctly
                    from openpyxl.styles import Alignment
                    existing = cell.alignment
                    cell.alignment = Alignment(
                        wrap_text=True,
                        horizontal=existing.horizontal,
                        vertical=existing.vertical,
                    )
                else:
                    cell.value = lrmx_dict.get(field, '') or None

        save_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(save_path)
        return ('updated' if update_only else 'created'), stem
