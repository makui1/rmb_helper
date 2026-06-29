from pathlib import Path
from typing import Callable, Optional
import openpyxl
from .lrmx import LrmxFile


class MatchMode:
    ID_CARD = 'ShenFenZheng'
    NAME = 'XingMing'
    NAME_AND_ID = 'both'


class ExcelHandler:
    def __init__(self, excel_path: Path, lrmx_files: list, match_mode: str) -> None:
        self.excel_path = Path(excel_path)
        self.lrmx_files = [Path(f) for f in lrmx_files]
        self.match_mode = match_mode

    def _make_key(self, lf: LrmxFile) -> str:
        """构建 lrmx 文件的匹配 key"""
        if self.match_mode == MatchMode.ID_CARD:
            return lf.get('ShenFenZheng').strip()
        if self.match_mode == MatchMode.NAME:
            return lf.get('XingMing').strip()
        return lf.get('XingMing').strip() + lf.get('ShenFenZheng').strip()

    def _excel_key(
        self,
        row: dict,
        id_col: str | None,
        name_col: str | None,
    ) -> str:
        """构建 Excel 行的匹配 key，与 _make_key 使用相同规则"""
        if self.match_mode == MatchMode.ID_CARD:
            return str(row.get(id_col) or '').strip() if id_col else ''
        if self.match_mode == MatchMode.NAME:
            return str(row.get(name_col) or '').strip() if name_col else ''
        id_val = str(row.get(id_col) or '').strip() if id_col else ''
        name_val = str(row.get(name_col) or '').strip() if name_col else ''
        return name_val + id_val

    def _load_index(self) -> dict[str, LrmxFile]:
        index: dict[str, LrmxFile] = {}
        for f in self.lrmx_files:
            try:
                lf = LrmxFile(f)
                key = self._make_key(lf)
                if key:
                    index[key] = lf
            except Exception:
                pass
        return index

    def update(
        self,
        field_mapping: dict[str, str],
        fields_to_write: list[str],
        header_row: int = 1,
        match_excel_col_for_id: str | None = None,
        match_excel_col_for_name: str | None = None,
        progress_cb: Optional[Callable[[str], None]] = None,
    ) -> list[str]:
        """
        从 Excel 读取数据，更新匹配的 lrmx 文件。

        field_mapping:           excel列名 → lrmx字段名
        fields_to_write:         实际要写入的 lrmx 字段名（field_mapping 值的子集）
        header_row:              Excel 表头行号（1-based）
        match_excel_col_for_id:  用于匹配身份证的 Excel 列名
        match_excel_col_for_name:用于匹配姓名的 Excel 列名
        """
        if self.match_mode == MatchMode.ID_CARD and not match_excel_col_for_id:
            raise ValueError('ID_CARD 匹配模式需要提供 match_excel_col_for_id')
        if self.match_mode == MatchMode.NAME and not match_excel_col_for_name:
            raise ValueError('NAME 匹配模式需要提供 match_excel_col_for_name')
        if self.match_mode == MatchMode.NAME_AND_ID and not (match_excel_col_for_id and match_excel_col_for_name):
            raise ValueError('NAME_AND_ID 匹配模式需要同时提供 match_excel_col_for_id 和 match_excel_col_for_name')
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active

        headers = [
            cell.value
            for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row))
        ]

        excel_index: dict[str, dict] = {}
        for row_values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row = dict(zip(headers, row_values))
            key = self._excel_key(row, match_excel_col_for_id, match_excel_col_for_name)
            if key:
                excel_index[key] = row

        lrmx_index = self._load_index()
        logs: list[str] = []

        for lrmx_key, lf in lrmx_index.items():
            name = lf.get('XingMing') or lrmx_key
            if lrmx_key not in excel_index:
                msg = f'△ {name}  未在名册中找到匹配记录'
                logs.append(msg)
                if progress_cb:
                    progress_cb(msg)
                continue

            excel_row = excel_index[lrmx_key]
            backup = lf.path.with_suffix('.lrmx.bak')
            lf.path.rename(backup)
            updated = 0
            try:
                for lrmx_field in fields_to_write:
                    excel_col = next(
                        (c for c, f in field_mapping.items() if f == lrmx_field),
                        None,
                    )
                    if excel_col and excel_col in excel_row:
                        val = excel_row[excel_col]
                        if val is not None:
                            lf.set(lrmx_field, str(val))
                            updated += 1
                lf.save(lf.path)
                msg = f'✓ {name}  已更新 {updated} 个字段'
                logs.append(msg)
                if progress_cb:
                    progress_cb(msg)
            except Exception as e:
                backup.rename(lf.path)
                msg = f'✗ {name}  {e}'
                logs.append(msg)
                if progress_cb:
                    progress_cb(msg)

        return logs

    def export_to_excel(
        self,
        field_mapping: dict[str, str],
        fields_to_write: list[str],
        converters: dict[str, str] | None = None,
        header_row: int = 1,
        match_excel_col_for_id: str | None = None,
        match_excel_col_for_name: str | None = None,
        progress_cb: Optional[Callable[[str], None]] = None,
    ) -> list[str]:
        """
        从 LRMX 文件读取数据，更新匹配的 Excel 行。

        field_mapping:           lrmx字段名 → excel列名 (与 update() 方向相反)
        fields_to_write:         实际要写入 Excel 的 lrmx 字段名
        converters:              lrmx字段名 → 转换器代码（仅需转换的字段）
        header_row:              Excel 表头行号（1-based）
        match_excel_col_for_id:  用于匹配身份证的 Excel 列名
        match_excel_col_for_name:用于匹配姓名的 Excel 列名
        """
        from app.core.converters import execute_converter

        converters = converters or {}

        if self.match_mode == MatchMode.ID_CARD and not match_excel_col_for_id:
            raise ValueError('ID_CARD 匹配模式需要提供 match_excel_col_for_id')
        if self.match_mode == MatchMode.NAME and not match_excel_col_for_name:
            raise ValueError('NAME 匹配模式需要提供 match_excel_col_for_name')
        if self.match_mode == MatchMode.NAME_AND_ID and not (
            match_excel_col_for_id and match_excel_col_for_name
        ):
            raise ValueError(
                'NAME_AND_ID 匹配模式需要同时提供 match_excel_col_for_id '
                '和 match_excel_col_for_name'
            )

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active

        headers = [
            cell.value
            for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row))
        ]

        excel_index: dict[str, tuple[int, dict]] = {}
        for row_idx, row_values in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
        ):
            row = dict(zip(headers, row_values))
            key = self._excel_key(row, match_excel_col_for_id, match_excel_col_for_name)
            if key:
                excel_index[key] = (row_idx, row)

        lrmx_index = self._load_index()
        logs: list[str] = []

        # 收集需要新增的列（Excel 表头中不存在的字段）
        existing_cols = set(headers)
        new_columns: list[str] = []
        for lrmx_field in fields_to_write:
            excel_col = field_mapping.get(lrmx_field, lrmx_field)
            if excel_col not in existing_cols:
                new_columns.append(excel_col)
                existing_cols.add(excel_col)

        # 构建 headers 查找索引
        header_index: dict[str, int] = {}
        for i, h in enumerate(headers):
            if h is not None:
                header_index[str(h)] = i + 1

        # 追加新列到末尾（必须在写数据之前，否则 header_index 缺少新列）
        last_col = ws.max_column
        for i, col_name in enumerate(new_columns):
            col_idx = last_col + 1 + i
            ws.cell(row=header_row, column=col_idx, value=col_name)
            header_index[col_name] = col_idx

        updated_count = 0

        for lrmx_key, lf in lrmx_index.items():
            name = lf.get('XingMing') or lrmx_key
            if lrmx_key not in excel_index:
                msg = f'△ {name}  未在名册中找到匹配记录'
                logs.append(msg)
                if progress_cb:
                    progress_cb(msg)
                continue

            row_idx, excel_row = excel_index[lrmx_key]
            updated = 0
            try:
                for lrmx_field in fields_to_write:
                    excel_col = field_mapping.get(lrmx_field, lrmx_field)
                    val = lf.get(lrmx_field)

                    # 应用转换器
                    converter_code = converters.get(lrmx_field, '')
                    if converter_code and val:
                        val = execute_converter(converter_code, val)

                    # 写入 Excel 单元格
                    col_idx = header_index.get(excel_col)
                    if col_idx is not None and val:
                        ws.cell(row=row_idx, column=col_idx, value=val)
                        updated += 1

                if updated > 0:
                    updated_count += 1
                msg = f'✓ {name}  已更新 {updated} 个字段'
                logs.append(msg)
                if progress_cb:
                    progress_cb(msg)
            except Exception as e:
                msg = f'✗ {name}  {e}'
                logs.append(msg)
                if progress_cb:
                    progress_cb(msg)

        # 备份并保存
        import shutil
        backup = self.excel_path.with_suffix('.xlsx.bak')
        shutil.copy2(self.excel_path, backup)
        try:
            wb.save(self.excel_path)
        except Exception:
            shutil.copy2(backup, self.excel_path)
            raise

        summary = (
            f'共处理 {len(lrmx_index)} 个文件，更新 {updated_count} 行'
        )
        if new_columns:
            summary += f'，新增 {len(new_columns)} 列'
        logs.insert(0, summary)
        return logs
