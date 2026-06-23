"""Batch verification: compare lrmx fields against an Excel roster."""
from __future__ import annotations

import difflib
import html as _html
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Optional

import openpyxl

from app.core.lrmx import LrmxFile
from app.core.excel_handler import MatchMode
from app.core.compare_rules import CompareRule, apply_rule

_INVIS = re.compile(
    r'[\x00-\x1f\x7f-\x9f'
    r'​-‏‪-‮⁠-⁤﻿'
    r'\s]'
)

_EXCLUDED_FIELDS = {'version'}
_CONTAINER_TAGS = {'JiaTingChengYuan'}

# ── static field registry ─────────────────────────────────────────────────────

LRMX_FIELDS: list[tuple[str, str]] = [
    ('XingMing',                              '姓名'),
    ('XingBie',                               '性别'),
    ('ChuShengNianYue',                       '出生年月'),
    ('MinZu',                                 '民族'),
    ('JiGuan',                                '籍贯'),
    ('ChuShengDi',                            '出生地'),
    ('RuDangShiJian',                         '入党时间'),
    ('CanJiaGongZuoShiJian',                  '参加工作时间'),
    ('DaoLingNianYue',                        '到龄年月'),
    ('JianKangZhuangKuang',                   '健康状况'),
    ('ZhuanYeJiShuZhiWu',                     '专业技术职务'),
    ('ShuXiZhuanYeYouHeZhuanChang',           '熟悉专业有何专长'),
    ('QuanRiZhiJiaoYu_XueLi',                '全日制学历'),
    ('QuanRiZhiJiaoYu_XueWei',               '全日制学位'),
    ('QuanRiZhiJiaoYu_XueLi_BiYeYuanXiaoXi', '全日制学历毕业院校'),
    ('QuanRiZhiJiaoYu_XueWei_BiYeYuanXiaoXi','全日制学位毕业院校'),
    ('ZaiZhiJiaoYu_XueLi',                   '在职学历'),
    ('ZaiZhiJiaoYu_XueWei',                  '在职学位'),
    ('ZaiZhiJiaoYu_XueLi_BiYeYuanXiaoXi',   '在职学历毕业院校'),
    ('ZaiZhiJiaoYu_XueWei_BiYeYuanXiaoXi',  '在职学位毕业院校'),
    ('XianRenZhiWu',                          '现任职务'),
    ('NiRenZhiWu',                            '拟任职务'),
    ('NiMianZhiWu',                           '拟免职务'),
    ('JianLi',                                '简历'),
    ('JiangChengQingKuang',                   '奖惩情况'),
    ('NianDuKaoHeJieGuo',                     '年度考核结果'),
    ('RenMianLiYou',                          '任免理由'),
    ('ChengBaoDanWei',                        '呈报单位'),
    ('GaiGeQianRenZhiNianLingJieXian',        '改革前任职年龄界限'),
    ('JiSuanNianLingShiJian',                 '计算年龄时间'),
    ('TianBiaoShiJian',                       '填表时间'),
    ('TianBiaoRen',                           '填表人'),
    ('ShenFenZheng',                          '身份证号'),
]

# Default Excel column aliases for each lrmx field.
# The verify tab uses these to auto-map when headers are loaded.
DEFAULT_FIELD_ALIASES: dict[str, list[str]] = {
    'XingMing':                              ['姓名', '名字', '干部姓名'],
    'XingBie':                               ['性别'],
    'ChuShengNianYue':                       ['出生年月', '出生日期', '出生年月日'],
    'MinZu':                                 ['民族'],
    'JiGuan':                                ['籍贯'],
    'ChuShengDi':                            ['出生地'],
    'ShenFenZheng':                          ['身份证号', '身份证', '证件号码', '证件号', '身份证号码'],
    'RuDangShiJian':                         ['入党时间', '入党日期', '入党年月'],
    'CanJiaGongZuoShiJian':                  ['参加工作时间', '参工时间', '参加工作日期', '参加工作年月'],
    'DaoLingNianYue':                        ['到龄年月', '到龄日期'],
    'JianKangZhuangKuang':                   ['健康状况', '健康'],
    'ZhuanYeJiShuZhiWu':                     ['专业技术职务', '技术职务', '专技职务'],
    'QuanRiZhiJiaoYu_XueLi':                ['全日制学历', '学历', '最高学历'],
    'QuanRiZhiJiaoYu_XueWei':               ['全日制学位', '学位', '最高学位'],
    'QuanRiZhiJiaoYu_XueLi_BiYeYuanXiaoXi': ['全日制学历毕业院校', '毕业院校'],
    'ZaiZhiJiaoYu_XueLi':                   ['在职学历'],
    'ZaiZhiJiaoYu_XueWei':                  ['在职学位'],
    'XianRenZhiWu':                          ['现任职务', '职务', '现职务'],
    'NiRenZhiWu':                            ['拟任职务'],
    'NiMianZhiWu':                           ['拟免职务'],
    'RenMianLiYou':                          ['任免理由'],
    'TianBiaoRen':                           ['填表人'],
    'ChengBaoDanWei':                        ['呈报单位', '单位'],
}


def _strip(s: str) -> str:
    return _INVIS.sub('', s)


def read_excel_headers(path: Path, header_row: int = 1) -> list[str]:
    """Return the list of column headers from the given row (1-indexed)."""
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if i == header_row:
            wb.close()
            return [str(c) if c is not None else '' for c in row]
    wb.close()
    return []


def char_diff_html(a: str, b: str) -> tuple[str, str]:
    """Return (a_html, b_html) with differing character spans highlighted."""
    matcher = difflib.SequenceMatcher(None, a, b, autojunk=False)
    ah: list[str] = []
    bh: list[str] = []
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        ea = _html.escape(a[i1:i2], quote=False)
        eb = _html.escape(b[j1:j2], quote=False)
        if tag == 'equal':
            ah.append(ea)
            bh.append(eb)
        elif tag == 'replace':
            ah.append(f'<span class="del">{ea}</span>')
            bh.append(f'<span class="ins">{eb}</span>')
        elif tag == 'delete':
            ah.append(f'<span class="del">{ea}</span>')
        elif tag == 'insert':
            bh.append(f'<span class="ins">{eb}</span>')
    return ''.join(ah), ''.join(bh)


@dataclass
class FieldResult:
    field: str
    excel_val: str      # raw (pre-strip) for display
    lrmx_val: str       # raw (pre-strip) for display
    match: bool


@dataclass
class PersonResult:
    name: str
    lrmx_path: str
    status: str         # 'ok' | 'diff' | 'not_found' | 'error'
    fields: list[FieldResult] = field(default_factory=list)
    error_msg: str = ''


class VerifyHandler:
    def __init__(
        self,
        excel_path: Path,
        lrmx_files: list,
        match_mode: str,
        header_row: int,
        field_mapping: dict[str, str],   # excel_col → lrmx_field
        match_excel_col_for_id: Optional[str],
        match_excel_col_for_name: Optional[str],
        compare_rules: dict[str, CompareRule] | None = None,
    ) -> None:
        self.excel_path = Path(excel_path)
        self.lrmx_files = [Path(f) for f in lrmx_files]
        self.match_mode = match_mode
        self.header_row = header_row
        self.field_mapping = field_mapping
        self._id_col = match_excel_col_for_id
        self._name_col = match_excel_col_for_name
        self.compare_rules: dict[str, CompareRule] = compare_rules or {}

    def _load_excel_index(self) -> dict[str, dict]:
        """Return {match_key: row_dict} from Excel."""
        wb = openpyxl.load_workbook(str(self.excel_path), read_only=True, data_only=True)
        ws = wb.active
        headers: list[str] = []
        index: dict[str, dict] = {}
        data_start = self.header_row + 1
        for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            if i == self.header_row:
                headers = [str(c) if c is not None else '' for c in row]
                continue
            if i < data_start:
                continue
            row_dict = dict(zip(headers, row))
            key = self._excel_key(row_dict)
            if key:
                index[key] = row_dict
        wb.close()
        return index

    def _excel_key(self, row: dict) -> str:
        name = _strip(str(row.get(self._name_col) or '')) if self._name_col else ''
        id_  = _strip(str(row.get(self._id_col)   or '')) if self._id_col   else ''
        if self.match_mode == MatchMode.ID_CARD:
            return id_
        if self.match_mode == MatchMode.NAME:
            return name
        return name + id_

    def _lrmx_key(self, lf: LrmxFile) -> str:
        name = _strip(lf.get('XingMing'))
        id_  = _strip(lf.get('ShenFenZheng'))
        if self.match_mode == MatchMode.ID_CARD:
            return id_
        if self.match_mode == MatchMode.NAME:
            return name
        return name + id_

    def verify(
        self,
        progress_cb: Optional[Callable[[PersonResult], None]] = None,
    ) -> list[PersonResult]:
        excel_index = self._load_excel_index()
        results: list[PersonResult] = []

        for lrmx_path in self.lrmx_files:
            try:
                lf = LrmxFile(lrmx_path)
            except Exception as e:
                r = PersonResult(
                    name=lrmx_path.stem,
                    lrmx_path=str(lrmx_path),
                    status='error',
                    error_msg=str(e),
                )
                results.append(r)
                if progress_cb:
                    progress_cb(r)
                continue

            name = lf.get('XingMing') or lrmx_path.stem
            key = self._lrmx_key(lf)

            if key not in excel_index:
                r = PersonResult(
                    name=name,
                    lrmx_path=str(lrmx_path),
                    status='not_found',
                )
                results.append(r)
                if progress_cb:
                    progress_cb(r)
                continue

            excel_row = excel_index[key]
            field_results: list[FieldResult] = []
            any_diff = False

            for excel_col, lrmx_field in self.field_mapping.items():
                excel_raw = str(excel_row.get(excel_col) or '')
                lrmx_raw  = lf.get(lrmx_field)
                match = _strip(excel_raw) == _strip(lrmx_raw)
                if not match:
                    rule = self.compare_rules.get(lrmx_field)
                    if rule:
                        match = apply_rule(rule, excel_raw, lrmx_raw)
                if not match:
                    any_diff = True
                field_results.append(FieldResult(
                    field=lrmx_field,
                    excel_val=excel_raw,
                    lrmx_val=lrmx_raw,
                    match=match,
                ))

            r = PersonResult(
                name=name,
                lrmx_path=str(lrmx_path),
                status='diff' if any_diff else 'ok',
                fields=field_results,
            )
            results.append(r)
            if progress_cb:
                progress_cb(r)

        return results
